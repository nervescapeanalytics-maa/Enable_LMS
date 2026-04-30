"""
ENABLE PROGRAM — Admin Utilities
Shared mixins, actions, and helpers for the enhanced admin console.
"""
import csv
import io
import json
from datetime import datetime, timedelta

from django.contrib import admin, messages
from django.http import HttpResponse, HttpResponseRedirect
from django.template.response import TemplateResponse
from django.urls import path, reverse
from django.utils import timezone
from django.utils.html import format_html
from django.db.models import Count, Q


# ═══════════════════════════════════════════════════════════════════
# EXPORT ACTIONS
# ═══════════════════════════════════════════════════════════════════

def export_as_csv(modeladmin, request, queryset):
    """Export selected records as CSV."""
    meta = modeladmin.model._meta
    field_names = [f.name for f in meta.fields]

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename={meta.verbose_name_plural.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

    writer = csv.writer(response)
    writer.writerow(field_names)

    for obj in queryset:
        row = []
        for field in field_names:
            value = getattr(obj, field)
            if isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            row.append(str(value) if value is not None else '')
        writer.writerow(row)

    modeladmin.message_user(request, f"Exported {queryset.count()} {meta.verbose_name_plural} to CSV.", messages.SUCCESS)
    return response

export_as_csv.short_description = "📥 Export selected as CSV"


def export_as_json(modeladmin, request, queryset):
    """Export selected records as JSON."""
    meta = modeladmin.model._meta
    field_names = [f.name for f in meta.fields]

    data = []
    for obj in queryset:
        row = {}
        for field in field_names:
            value = getattr(obj, field)
            if isinstance(value, datetime):
                value = value.isoformat()
            elif hasattr(value, 'pk'):
                value = str(value.pk)
            else:
                value = str(value) if value is not None else None
            row[field] = value
        data.append(row)

    response = HttpResponse(
        json.dumps(data, indent=2, ensure_ascii=False),
        content_type='application/json'
    )
    response['Content-Disposition'] = f'attachment; filename={meta.verbose_name_plural.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'

    modeladmin.message_user(request, f"Exported {queryset.count()} {meta.verbose_name_plural} to JSON.", messages.SUCCESS)
    return response

export_as_json.short_description = "📥 Export selected as JSON"


# ═══════════════════════════════════════════════════════════════════
# STATUS ACTIONS
# ═══════════════════════════════════════════════════════════════════

def activate_selected(modeladmin, request, queryset):
    """Activate selected records (set status=ACTIVE or is_active=True)."""
    count = 0
    if hasattr(modeladmin.model, 'status'):
        count = queryset.update(status='ACTIVE')
    elif hasattr(modeladmin.model, 'is_active'):
        count = queryset.update(is_active=True)
    modeladmin.message_user(request, f"Activated {count} records.", messages.SUCCESS)

activate_selected.short_description = "✅ Activate selected"


def deactivate_selected(modeladmin, request, queryset):
    """Deactivate selected records."""
    count = 0
    if hasattr(modeladmin.model, 'status'):
        count = queryset.update(status='INACTIVE')
    elif hasattr(modeladmin.model, 'is_active'):
        count = queryset.update(is_active=False)
    modeladmin.message_user(request, f"Deactivated {count} records.", messages.SUCCESS)

deactivate_selected.short_description = "🚫 Deactivate selected"


def suspend_selected(modeladmin, request, queryset):
    """Suspend selected records."""
    if hasattr(modeladmin.model, 'status'):
        count = queryset.update(status='SUSPENDED')
        modeladmin.message_user(request, f"Suspended {count} records.", messages.WARNING)

suspend_selected.short_description = "⏸ Suspend selected"


# ═══════════════════════════════════════════════════════════════════
# DISPLAY HELPERS
# ═══════════════════════════════════════════════════════════════════

def colored_status(status):
    """Returns HTML for a color-coded status badge with light backgrounds."""
    color_map = {
        'ACTIVE': ('#15803d', 'rgba(16, 185, 129, 0.15)'),
        'INACTIVE': ('#dc2626', 'rgba(239, 68, 68, 0.15)'),
        'SUSPENDED': ('#d97706', 'rgba(245, 158, 11, 0.15)'),
        'PENDING_VERIFICATION': ('#2563eb', 'rgba(59, 130, 246, 0.15)'),
        'PENDING': ('#2563eb', 'rgba(59, 130, 246, 0.15)'),
        'DRAFT': ('#64748b', 'rgba(148, 163, 184, 0.15)'),
        'PUBLISHED': ('#15803d', 'rgba(16, 185, 129, 0.15)'),
        'COMPLETED': ('#15803d', 'rgba(16, 185, 129, 0.15)'),
        'CANCELLED': ('#dc2626', 'rgba(239, 68, 68, 0.15)'),
        'IN_PROGRESS': ('#d97706', 'rgba(245, 158, 11, 0.15)'),
        'SCHEDULED': ('#7c3aed', 'rgba(139, 92, 246, 0.15)'),
        'LIVE': ('#dc2626', 'rgba(239, 68, 68, 0.15)'),
        'OPEN': ('#15803d', 'rgba(16, 185, 129, 0.15)'),
        'CLOSED': ('#64748b', 'rgba(148, 163, 184, 0.15)'),
        'RESOLVED': ('#15803d', 'rgba(16, 185, 129, 0.15)'),
        'PAID': ('#15803d', 'rgba(16, 185, 129, 0.15)'),
        'UNPAID': ('#dc2626', 'rgba(239, 68, 68, 0.15)'),
        'PARTIAL': ('#d97706', 'rgba(245, 158, 11, 0.15)'),
        True: ('#15803d', 'rgba(16, 185, 129, 0.15)'),
        False: ('#dc2626', 'rgba(239, 68, 68, 0.15)'),
    }
    fg, bg = color_map.get(status, ('#64748b', 'rgba(148, 163, 184, 0.15)'))
    label = str(status).replace('_', ' ').title() if isinstance(status, str) else ('Active' if status else 'Inactive')
    return format_html(
        '<span style="display:inline-flex;align-items:center;padding:4px 12px;'
        'border-radius:6px;font-size:0.72rem;font-weight:600;'
        'color:{};background:{};letter-spacing:0.03em;'
        'border:1px solid {};">{}</span>',
        fg, bg, fg, label
    )


# ═══════════════════════════════════════════════════════════════════
# ENHANCED BASE ADMIN
# ═══════════════════════════════════════════════════════════════════

class EnhancedModelAdmin(admin.ModelAdmin):
    """Base ModelAdmin with export actions, colored statuses, and enhanced UI."""

    actions = [export_as_csv, export_as_json, activate_selected, deactivate_selected]
    list_per_page = 25
    show_full_result_count = True
    save_on_top = True

    class Media:
        css = {'all': ('css/admin_theme.css',)}
        js = ('js/column_manager.js',)

    def _model_has_tenant(self):
        return any(f.name == 'tenant' for f in self.model._meta.fields)

    def _get_session_tenant(self, request):
        """Return the tenant selected in the admin session switcher, or None."""
        tid = request.session.get('admin_tenant_id')
        if tid:
            from tenants.models import Tenant
            try:
                return Tenant.objects.get(id=tid)
            except Tenant.DoesNotExist:
                pass
        return None

    def lookup_allowed(self, lookup, value):
        """Allow __gte / __lt / __lte lookups for date range filtering."""
        if any(lookup.endswith(s) for s in ('__gte', '__lt', '__lte')):
            field_name = lookup.rsplit('__', 1)[0]
            try:
                self.model._meta.get_field(field_name)
                return True
            except Exception:
                pass
        return super().lookup_allowed(lookup, value)

    def save_model(self, request, obj, form, change):
        if not change and self._model_has_tenant() and not obj.tenant_id:
            tenant = self._get_session_tenant(request)
            if not tenant:
                from tenants.models import Tenant
                tenant = Tenant.objects.filter(status__in=['ACTIVE', 'active']).first()
            obj.tenant = tenant
        super().save_model(request, obj, form, change)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        tenant_id = request.session.get('admin_tenant_id')
        if tenant_id and self._model_has_tenant():
            qs = qs.filter(tenant_id=tenant_id)
        return qs

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == 'tenant':
            tenant = self._get_session_tenant(request)
            if tenant:
                kwargs['initial'] = tenant
                from tenants.models import Tenant
                kwargs['queryset'] = Tenant.objects.filter(id=tenant.id)
            else:
                from tenants.models import Tenant
                kwargs['initial'] = Tenant.objects.filter(status__in=['ACTIVE', 'active']).first()
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    # ── Per-column → model-field map for server-side filtering ──
    # Subclasses may override; otherwise auto-derived from list_filter + admin_order_field.
    column_filter_fields = None  # type: dict | None

    # ── Section-level RBAC ──
    # When a non-superuser has a UserStaffRole, only StaffRolePermission rows
    # for that role grant access. has_module_permission also defers to the same
    # check so unauthorized sections disappear from the admin index.
    def _staff_role_permission(self, request):
        """Return StaffRolePermission for current user / this model, or None."""
        user = getattr(request, 'user', None)
        if user is None or not user.is_authenticated:
            return None
        if user.is_superuser:
            return 'SUPERUSER'  # sentinel: bypass all checks
        try:
            from accounts.models import UserStaffRole, StaffRolePermission
        except Exception:
            return None
        try:
            link = UserStaffRole.objects.select_related('role').get(user=user, is_active=True)
        except Exception:
            # No staff-role assignment → fall back to legacy behaviour
            return 'LEGACY'
        try:
            return StaffRolePermission.objects.get(
                role=link.role,
                app_label=self.opts.app_label,
                model_name=self.opts.model_name,
            )
        except Exception:
            return False  # role assigned but no perms for this model → DENY

    def _check_perm(self, request, action):
        sentinel = self._staff_role_permission(request)
        if sentinel == 'SUPERUSER' or sentinel == 'LEGACY':
            return None  # let parent handle
        if sentinel is False or sentinel is None:
            return False
        return getattr(sentinel, f'can_{action}', False)

    def has_view_permission(self, request, obj=None):
        r = self._check_perm(request, 'view')
        if r is None:
            return super().has_view_permission(request, obj)
        return bool(r)

    def has_add_permission(self, request):
        r = self._check_perm(request, 'add')
        if r is None:
            return super().has_add_permission(request)
        return bool(r)

    def has_change_permission(self, request, obj=None):
        r = self._check_perm(request, 'change')
        if r is None:
            return super().has_change_permission(request, obj)
        return bool(r)

    def has_delete_permission(self, request, obj=None):
        r = self._check_perm(request, 'delete')
        if r is None:
            return super().has_delete_permission(request, obj)
        return bool(r)

    def has_module_permission(self, request):
        # Allow module index if the user has any view perm for any of its models.
        sentinel = self._staff_role_permission(request)
        if sentinel == 'SUPERUSER' or sentinel == 'LEGACY':
            return super().has_module_permission(request)
        if sentinel is False or sentinel is None:
            return False
        return bool(getattr(sentinel, 'can_view', False))

    def get_column_filter_fields(self):
        """Map list_display column index → underlying model field name.

        Every list_display column whose underlying value is a real model field
        (CharField, ForeignKey, choices, BooleanField, …) becomes filterable.
        Date/Time/JSON/binary fields are excluded because their distinct-value
        dropdown isn't useful (use the sidebar date filter instead).
        The result is consumed by the Excel-like column dropdown in the changelist.
        """
        if self.column_filter_fields is not None:
            return dict(self.column_filter_fields)

        list_display = list(getattr(self, 'list_display', []) or [])

        try:
            field_objs = {f.name: f for f in self.model._meta.get_fields() if hasattr(f, 'name')}
        except Exception:
            field_objs = {}

        from django.db import models as dj_models
        skip_types = (
            dj_models.DateField, dj_models.DateTimeField, dj_models.TimeField,
            dj_models.JSONField, dj_models.BinaryField,
            dj_models.FileField, dj_models.ImageField,
        )

        mapping = {}
        # Column 0 is the action checkbox; list_display starts at column 1.
        for idx, col in enumerate(list_display, start=1):
            field_name = None
            if isinstance(col, str):
                if col in field_objs:
                    field_name = col
                else:
                    callable_obj = getattr(self, col, None) or getattr(self.model, col, None)
                    field_name = getattr(callable_obj, 'admin_order_field', None)
            else:
                field_name = getattr(col, 'admin_order_field', None)

            if not field_name:
                continue
            # Strip any "-" sort prefix and ".."-style joins (e.g. parent__name)
            base_field_name = field_name.lstrip('-').split('__', 1)[0]
            f = field_objs.get(base_field_name)
            if f is None:
                continue
            if isinstance(f, skip_types):
                continue
            mapping[idx] = base_field_name
        return mapping

    def changelist_view(self, request, extra_context=None):
        """Override to support dynamic list_per_page + inject filter-fields map.

        Uses a request-local override (not a class-level mutation) so per-page
        choices don't leak between requests / workers.
        """
        per_page = request.GET.get('list_per_page')
        per_page_override = None
        if per_page:
            try:
                per_page_val = int(per_page)
                if per_page_val in (10, 20, 25, 50, 100, 200):
                    per_page_override = per_page_val
            except (ValueError, TypeError):
                pass
            # Remove list_per_page from GET so Django doesn't treat it as a filter
            modified_get = request.GET.copy()
            modified_get.pop('list_per_page', None)
            request.GET = modified_get

        # Stash on the request so get_changelist_instance / templates can read it
        # without us having to mutate the shared ModelAdmin instance.
        if per_page_override is not None:
            request._enf_list_per_page = per_page_override

        extra_context = extra_context or {}
        try:
            extra_context['enf_filter_fields_json'] = json.dumps(self.get_column_filter_fields())
        except Exception:
            extra_context['enf_filter_fields_json'] = '{}'
        # Class-level toggle to hide the "Tools" command palette button.
        if getattr(self, 'enf_hide_tools', False):
            extra_context.setdefault('enf_disable_tools', True)
        return super().changelist_view(request, extra_context=extra_context)

    def get_changelist_instance(self, request):
        """Apply the request-local list_per_page override before ChangeList is built."""
        override = getattr(request, '_enf_list_per_page', None)
        if override is not None:
            original = self.list_per_page
            self.list_per_page = override
            try:
                return super().get_changelist_instance(request)
            finally:
                # Restore so subsequent requests on this worker see the original.
                self.list_per_page = original
        return super().get_changelist_instance(request)

    def filter_choices_view(self, request):
        """Return distinct values for a model field across the *full* (filtered) queryset.

        Used by the column header dropdown so filter values reflect *all* rows,
        not just the current page. Honors other active filters except for the
        field being queried (so user sees all available choices for that column).

        Allowed fields are derived from `get_column_filter_fields()` which now
        covers every column in `list_display` whose underlying value is a real
        model field — not just those declared in `list_filter`.
        """
        from django.http import JsonResponse
        field = request.GET.get('field', '').strip()
        if not field:
            return JsonResponse({'error': 'field required'}, status=400)

        # Whitelist: only fields exposed via the column filter map are queryable.
        allowed = set(self.get_column_filter_fields().values())
        # Also allow fields explicitly declared in list_filter.
        for f in (getattr(self, 'list_filter', []) or []):
            if isinstance(f, str):
                allowed.add(f)
            elif isinstance(f, (list, tuple)) and f:
                allowed.add(f[0])
        if field not in allowed:
            return JsonResponse({'error': 'field not filterable'}, status=403)

        try:
            field_obj = self.model._meta.get_field(field)
        except Exception:
            return JsonResponse({'error': 'unknown field'}, status=400)

        qs = self.get_queryset(request)
        # Apply OTHER active filters (so dropdown values are contextual).
        for k, v in request.GET.items():
            if k in (field, 'p', 'list_per_page', 'q', 'o', 'all', '_filter_field'):
                continue
            try:
                qs = qs.filter(**{k: v})
            except Exception:
                continue

        # ── ForeignKey: return {id, label, count} so the UI can navigate by id ──
        if field_obj.is_relation and field_obj.many_to_one:
            label_attr = '__str__'
            rel_model = field_obj.related_model
            # Try common name fields for a friendlier label
            for candidate in ('name', 'title', 'code', 'label'):
                try:
                    rel_model._meta.get_field(candidate)
                    label_attr = candidate
                    break
                except Exception:
                    continue

            rows = (
                qs.exclude(**{f'{field}__isnull': True})
                .values(field, f'{field}__{label_attr}' if label_attr != '__str__' else field)
                .annotate(c=Count('id'))
                .order_by(f'{field}__{label_attr}' if label_attr != '__str__' else field)[:2000]
            )
            choices = []
            seen = set()
            for r in rows:
                fk_id = r[field]
                if fk_id is None or fk_id in seen:
                    continue
                seen.add(fk_id)
                lbl = r.get(f'{field}__{label_attr}') if label_attr != '__str__' else None
                if not lbl:
                    try:
                        lbl = str(rel_model.objects.get(pk=fk_id))
                    except Exception:
                        lbl = str(fk_id)
                choices.append({'value': str(fk_id), 'label': str(lbl), 'count': r['c']})
            return JsonResponse({'field': field, 'choices': choices, 'current': request.GET.get(field, '')})

        # ── BooleanField: render Yes/No ──
        from django.db import models as dj_models
        if isinstance(field_obj, dj_models.BooleanField):
            rows = (
                qs.values(field)
                .annotate(c=Count('id'))
                .order_by(field)
            )
            label_map = {True: 'Yes', False: 'No', 'True': 'Yes', 'False': 'No', 1: 'Yes', 0: 'No'}
            choices = []
            for r in rows:
                if r[field] is None:
                    continue
                choices.append({
                    'value': str(r[field]),
                    'label': label_map.get(r[field], str(r[field])),
                    'count': r['c'],
                })
            return JsonResponse({'field': field, 'choices': choices, 'current': request.GET.get(field, '')})

        # ── Choices field: prefer human-readable display ──
        choices_def = getattr(field_obj, 'choices', None)
        rows = (
            qs.values(field)
            .annotate(c=Count('id'))
            .order_by('-c', field)[:2000]
        )
        if choices_def:
            label_map = dict(choices_def)
            choices = []
            for r in rows:
                if r[field] in (None, ''):
                    continue
                v = r[field]
                choices.append({
                    'value': str(v),
                    'label': str(label_map.get(v, v)),
                    'count': r['c'],
                })
            return JsonResponse({'field': field, 'choices': choices, 'current': request.GET.get(field, '')})

        # ── Plain CharField/TextField/IntegerField: value == label ──
        choices = [
            {'value': str(r[field]), 'label': str(r[field]), 'count': r['c']}
            for r in rows
            if r[field] not in (None, '')
        ]
        return JsonResponse({'field': field, 'choices': choices, 'current': request.GET.get(field, '')})

    def get_urls(self):
        urls = super().get_urls()
        info = (self.model._meta.app_label, self.model._meta.model_name)
        custom = [
            path(
                '_filter-choices/',
                self.admin_site.admin_view(self.filter_choices_view),
                name='%s_%s_filter_choices' % info,
            ),
        ]
        return custom + urls

    def status_badge(self, obj):
        status = getattr(obj, 'status', None)
        if status is None:
            status = getattr(obj, 'is_active', None)
        if status is not None:
            return colored_status(status)
        return '-'
    status_badge.short_description = 'Status'
    status_badge.admin_order_field = 'status'

    def created_display(self, obj):
        dt = getattr(obj, 'created_at', None)
        if dt:
            return format_html(
                '<span style="color:#94a3b8;font-size:0.82rem;" title="{}">{}</span>',
                dt.strftime('%Y-%m-%d %H:%M:%S'),
                dt.strftime('%d %b %Y')
            )
        return '-'
    created_display.short_description = 'Created'
    created_display.admin_order_field = 'created_at'

    def updated_display(self, obj):
        dt = getattr(obj, 'updated_at', None)
        if dt:
            return format_html(
                '<span style="color:#94a3b8;font-size:0.82rem;" title="{}">{}</span>',
                dt.strftime('%Y-%m-%d %H:%M:%S'),
                dt.strftime('%d %b %Y')
            )
        return '-'
    updated_display.short_description = 'Updated'
    updated_display.admin_order_field = 'updated_at'


# ═══════════════════════════════════════════════════════════════════
# IMPORT MIXIN
# ═══════════════════════════════════════════════════════════════════

class ImportExportMixin:
    """Adds import/export URL endpoints to a ModelAdmin."""
    change_list_template = 'admin/enhanced_change_list.html'

    # Override on a subclass to add custom header → field-name mappings.
    # Keys are case/whitespace-insensitive (normalized via _norm_header).
    # Example for TeacherAdmin:
    #     import_field_aliases = {
    #         'state': 'teacher_state',
    #         'city': 'teacher_city',
    #         'village': 'teacher_city',
    #         'district': 'teacher_district',
    #         'joining date': 'joining_date',
    #     }
    import_field_aliases: dict = {}

    @staticmethod
    def _norm_header(h):
        """Lowercase + collapse whitespace + strip non-alnum (keep spaces)."""
        if h is None:
            return ''
        s = str(h).strip().lower()
        # collapse runs of whitespace/underscores/hyphens to single space
        out = []
        prev_space = False
        for ch in s:
            if ch.isalnum():
                out.append(ch)
                prev_space = False
            elif ch in (' ', '_', '-', '/', '.'):
                if not prev_space and out:
                    out.append(' ')
                prev_space = True
        return ''.join(out).strip()

    def _build_alias_map(self):
        """Combine user-supplied aliases with auto-aliases derived from fields.

        Auto-alias rules:
          - field "first_name"  ← "first name", "firstname"
          - field "teacher_state" ← "teacher state"
        Plus the explicit map in `self.import_field_aliases` wins.
        """
        valid_fields = {f.name for f in self.model._meta.fields}
        amap = {}
        # Auto: spaces/underscores variants of the actual field name
        for f in valid_fields:
            amap[self._norm_header(f)] = f
            amap[self._norm_header(f.replace('_', ''))] = f
        # User overrides
        for k, v in (self.import_field_aliases or {}).items():
            amap[self._norm_header(k)] = v
        return amap, valid_fields

    def _normalize_row(self, raw_row):
        """Map raw header → field name; drop empty/unknown columns; strip values.

        Returns dict suitable for `Model.objects.create(**...)`.
        """
        amap, valid_fields = self._build_alias_map()
        out = {}
        for k, v in (raw_row or {}).items():
            if k is None:
                continue
            field = amap.get(self._norm_header(k))
            if not field or field not in valid_fields:
                continue
            if field in ('id', 'pk', 'uuid', 'created_at', 'updated_at'):
                continue
            if isinstance(v, str):
                v = v.strip()
            if v in (None, ''):
                # leave blank values out so model defaults apply
                continue
            out[field] = v
        return out

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-csv/', self.import_csv_view, name=f'{self.model._meta.app_label}_{self.model._meta.model_name}_import_csv'),
            path('import-excel/', self.import_excel_view, name=f'{self.model._meta.app_label}_{self.model._meta.model_name}_import_excel'),
            path('export-all-csv/', self.export_all_csv_view, name=f'{self.model._meta.app_label}_{self.model._meta.model_name}_export_all_csv'),
            path('export-all-excel/', self.export_all_excel_view, name=f'{self.model._meta.app_label}_{self.model._meta.model_name}_export_all_excel'),
            path('export-filtered-excel/', self.export_filtered_excel_view, name=f'{self.model._meta.app_label}_{self.model._meta.model_name}_export_filtered_excel'),
        ]
        return custom_urls + urls

    def import_csv_view(self, request):
        if request.method == 'POST' and request.FILES.get('csv_file'):
            csv_file = request.FILES['csv_file']
            try:
                decoded = csv_file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded))
                created_count = 0
                error_count = 0
                error_details = []
                for row_num, row in enumerate(reader, start=2):
                    try:
                        # Normalize headers (e.g. "State" → "teacher_state") and drop empties
                        filtered_row = self._normalize_row(row)
                        self.model.objects.create(**filtered_row)
                        created_count += 1
                    except Exception as e:
                        error_count += 1
                        if len(error_details) < 10:
                            error_details.append(f"Row {row_num}: {str(e)[:200]}")

                detail_msg = f"Import complete: {created_count} created, {error_count} errors."
                if error_details:
                    detail_msg += " First errors: " + " | ".join(error_details)

                self.message_user(
                    request,
                    detail_msg,
                    messages.SUCCESS if error_count == 0 else messages.WARNING
                )
            except Exception as e:
                self.message_user(request, f"Import failed: {str(e)}", messages.ERROR)

            return HttpResponseRedirect(request.META.get('HTTP_REFERER', '../'))

        context = {
            **self.admin_site.each_context(request),
            'title': f'Import {self.model._meta.verbose_name_plural.title()}',
            'model_name': self.model._meta.verbose_name_plural.title(),
            'opts': self.model._meta,
        }
        return TemplateResponse(request, 'admin/import_csv.html', context)

    def import_excel_view(self, request):
        if request.method == 'POST' and request.FILES.get('excel_file'):
            excel_file = request.FILES['excel_file']
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filename=io.BytesIO(excel_file.read()), read_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    self.message_user(request, "Excel file is empty or has no data rows.", messages.WARNING)
                    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '../'))

                headers = [str(h).strip() if h else '' for h in rows[0]]
                created_count = 0
                error_count = 0
                error_details = []

                for row_num, row in enumerate(rows[1:], start=2):
                    try:
                        clean_row = {}
                        for idx, header in enumerate(headers):
                            if header and idx < len(row):
                                val = row[idx]
                                clean_row[header] = str(val).strip() if val is not None else ''
                        # Normalize headers (e.g. "Joining Date" → "joining_date") and drop empties
                        filtered_row = self._normalize_row(clean_row)
                        self.model.objects.create(**filtered_row)
                        created_count += 1
                    except Exception as e:
                        error_count += 1
                        if len(error_details) < 10:
                            error_details.append(f"Row {row_num}: {str(e)[:200]}")

                wb.close()
                detail_msg = f"Import complete: {created_count} created, {error_count} errors."
                if error_details:
                    detail_msg += " First errors: " + " | ".join(error_details)
                self.message_user(
                    request,
                    detail_msg,
                    messages.SUCCESS if error_count == 0 else messages.WARNING
                )
            except Exception as e:
                self.message_user(request, f"Import failed: {str(e)}", messages.ERROR)

            return HttpResponseRedirect(request.META.get('HTTP_REFERER', '../'))

        context = {
            **self.admin_site.each_context(request),
            'title': f'Import {self.model._meta.verbose_name_plural.title()}',
            'model_name': self.model._meta.verbose_name_plural.title(),
            'opts': self.model._meta,
        }
        return TemplateResponse(request, 'admin/import_excel.html', context)

    def export_all_csv_view(self, request):
        queryset = self.model.objects.all()
        return export_as_csv(self, request, queryset)

    def export_all_excel_view(self, request):
        """Export all records as Excel (.xlsx)."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            self.message_user(request, "openpyxl is required for Excel export. Please install it.", messages.ERROR)
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', '../'))

        meta = self.model._meta
        queryset = self.model.objects.all()
        field_names = [f.name for f in meta.fields]

        wb = Workbook()
        ws = wb.active
        ws.title = meta.verbose_name_plural.title()[:31]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="844FC1", end_color="844FC1", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = []
        for name in field_names:
            try:
                field = meta.get_field(name)
                headers.append(field.verbose_name.title() if hasattr(field, 'verbose_name') else name)
            except Exception:
                headers.append(name.replace('_', ' ').title())

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_num, obj in enumerate(queryset, 2):
            for col_num, name in enumerate(field_names, 1):
                value = getattr(obj, name, '')
                if hasattr(value, '__call__'):
                    value = value()
                if value is None:
                    value = ''
                elif hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d %H:%M:%S') if hasattr(value, 'hour') else value.strftime('%Y-%m-%d')
                else:
                    value = str(value) if value else ''
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'{meta.verbose_name_plural.replace(" ", "_")}_all_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)

        self.message_user(request, f"Exported {queryset.count()} {meta.verbose_name_plural} to Excel.", messages.SUCCESS)
        return response

    def export_filtered_excel_view(self, request):
        """Export currently filtered/displayed records as Excel (.xlsx)."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            self.message_user(request, "openpyxl is required for Excel export. Please install it.", messages.ERROR)
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', '../'))

        from django.contrib.admin.views.main import ChangeList

        # Get the filtered queryset via ChangeList (same as Django's changelist_view)
        list_display = self.get_list_display(request)
        list_display_links = self.get_list_display_links(request, list_display)
        list_filter = self.get_list_filter(request)
        search_fields = self.get_search_fields(request)
        list_select_related = self.get_list_select_related(request)

        try:
            cl = ChangeList(
                request,
                self.model,
                list_display,
                list_display_links,
                list_filter,
                self.date_hierarchy,
                search_fields,
                list_select_related,
                self.list_per_page,
                self.list_max_show_all,
                self.list_editable,
                self,
                self.sortable_by,
                self.search_help_text,
            )
            # get_queryset applies search, filters, and ordering — NOT pagination
            queryset = cl.get_queryset(request)
        except Exception:
            # Fallback: manually apply filters
            queryset = self.get_queryset(request)
            search_term = request.GET.get('q', '').strip()
            if search_term and self.search_fields:
                search_query = Q()
                for field in self.search_fields:
                    search_query |= Q(**{f"{field}__icontains": search_term})
                queryset = queryset.filter(search_query)
            for key, value in request.GET.items():
                if key in ('q', 'p', 'o', 'list_per_page', '_changelist_filters', 'e', 'all'):
                    continue
                if value:
                    field_name = key.replace('__exact', '').replace('__in', '').replace('__isnull', '')
                    try:
                        self.model._meta.get_field(field_name)
                        queryset = queryset.filter(**{key: value})
                    except Exception:
                        pass

        meta = self.model._meta

        # Get field names from list_display
        field_names = []
        method_names = []
        for name in list_display:
            if hasattr(self.model, name):
                field_names.append(name)
            elif hasattr(self, name):
                method_names.append(name)
            else:
                try:
                    self.model._meta.get_field(name)
                    field_names.append(name)
                except Exception:
                    pass

        if not field_names:
            field_names = [f.name for f in meta.fields]

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = meta.verbose_name_plural.title()[:31]

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="844FC1", end_color="844FC1", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        headers = []
        for name in field_names:
            try:
                field = meta.get_field(name)
                headers.append(field.verbose_name.title() if hasattr(field, 'verbose_name') else name)
            except Exception:
                headers.append(name.replace('_', ' ').title())
        for name in method_names:
            method = getattr(self, name, None)
            if hasattr(method, 'short_description'):
                headers.append(method.short_description)
            else:
                headers.append(name.replace('_', ' ').title())

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data rows
        for row_num, obj in enumerate(queryset, 2):
            col_num = 1
            for name in field_names:
                value = getattr(obj, name, '')
                if hasattr(value, '__call__'):
                    value = value()
                if value is None:
                    value = ''
                elif hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d %H:%M:%S') if hasattr(value, 'hour') else value.strftime('%Y-%m-%d')
                else:
                    value = str(value) if value else ''
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                col_num += 1

            for name in method_names:
                method = getattr(self, name, None)
                if method:
                    try:
                        from django.utils.html import strip_tags
                        value = strip_tags(str(method(obj)))
                    except Exception:
                        value = ''
                else:
                    value = ''
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                col_num += 1

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'{meta.verbose_name_plural.replace(" ", "_")}_filtered_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)

        self.message_user(request, f"Exported {queryset.count()} filtered {meta.verbose_name_plural} to Excel.", messages.SUCCESS)
        return response
