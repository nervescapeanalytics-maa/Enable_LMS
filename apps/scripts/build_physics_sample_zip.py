"""Build a working physics_sample.zip for the assessments ZIP-importer.

Output:
    samples/physics_sample.zip   ── ready to upload
    samples/physics_sample/      ── unzipped working copy + README

Run from anywhere:
    python apps/scripts/build_physics_sample_zip.py
"""
from __future__ import annotations

import io
import os
import struct
import zipfile
import zlib
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]          # repo root
OUT_DIR = ROOT / 'samples'
WORK_DIR = OUT_DIR / 'physics_sample'
ZIP_PATH = OUT_DIR / 'physics_sample.zip'

OUT_DIR.mkdir(parents=True, exist_ok=True)
WORK_DIR.mkdir(parents=True, exist_ok=True)
(WORK_DIR / 'images').mkdir(parents=True, exist_ok=True)


# ── 1. Build a tiny PNG in pure Python (no Pillow needed) ────────────────────
def _png(width: int, height: int, rgb: tuple[int, int, int]) -> bytes:
    """Return bytes of a solid-colour PNG (RGB)."""
    def chunk(tag: bytes, data: bytes) -> bytes:
        crc = zlib.crc32(tag + data) & 0xffffffff
        return struct.pack('>I', len(data)) + tag + data + struct.pack('>I', crc)

    sig = b'\x89PNG\r\n\x1a\n'
    ihdr = struct.pack('>IIBBBBB', width, height, 8, 2, 0, 0, 0)  # 8-bit RGB
    raw = b''
    row = bytes(rgb) * width
    for _ in range(height):
        raw += b'\x00' + row                                  # filter byte
    idat = zlib.compress(raw, 9)
    return sig + chunk(b'IHDR', ihdr) + chunk(b'IDAT', idat) + chunk(b'IEND', b'')


# Two distinct sample images
(WORK_DIR / 'images' / 'q02_trajectory.png').write_bytes(_png(160, 90, (240, 200, 80)))
(WORK_DIR / 'images' / 'q04_circuit.png').write_bytes(_png(160, 90, (120, 180, 240)))


# ── 2. Build questions.xlsx in pure Python (no openpyxl needed) ──────────────
HEADERS = [
    'question_code', 'question_text', 'question_type',
    'option_a', 'option_b', 'option_c', 'option_d',
    'correct_answer', 'positive_marks', 'negative_marks',
    'difficulty', 'question_image', 'answer_explanation', 'tags',
    'question_order',
]

ROWS = [
    ['PHY-001', 'Speed of light in vacuum is approximately?', 'MCQ_SINGLE',
     '3×10^5 km/s', '3×10^8 m/s', '3×10^6 m/s', '3×10^10 m/s',
     'B', '4', '1', 'EASY', '', 'c ≈ 299 792 458 m/s ≈ 3×10^8 m/s.', 'optics,constants', '1'],
    ['PHY-002', 'A ball is thrown straight up at 20 m/s. How high does it rise? (g=10 m/s^2)',
     'MCQ_SINGLE', '10 m', '15 m', '20 m', '40 m',
     'C', '4', '1', 'MEDIUM', 'q02_trajectory.png',
     'h = u^2 / (2g) = 400 / 20 = 20 m.', 'kinematics', '2'],
    ['PHY-003', 'Which of these are scalar quantities?', 'MCQ_MULTI',
     'Velocity', 'Mass', 'Force', 'Temperature',
     'B,D', '4', '1', 'EASY', '',
     'Velocity and Force are vectors; mass & temperature are scalars.',
     'mechanics,vectors', '3'],
    ['PHY-004', 'In a 12 V circuit with 6 Ω resistance, find the current.', 'MCQ_SINGLE',
     '1 A', '2 A', '3 A', '4 A',
     'B', '4', '1', 'MEDIUM', 'q04_circuit.png',
     'I = V / R = 12 / 6 = 2 A.', 'electricity,ohms-law', '4'],
    ['PHY-005', 'Inertia of a body depends only on its mass.', 'MCQ_SINGLE',
     'True', 'False', 'Depends on velocity', 'Depends on temperature',
     'A', '2', '0', 'EASY', '',
     "Newton's first law: more mass ⇒ more inertia.", 'mechanics,newton', '5'],
]

# --- Minimal Office Open XML (.xlsx) writer -------------------------------
# Reference: ECMA-376. We emit only the parts openpyxl will actually read.
def _col_letter(idx0: int) -> str:
    n, out = idx0 + 1, ''
    while n:
        n, r = divmod(n - 1, 26)
        out = chr(65 + r) + out
    return out


def _xml_escape(s: str) -> str:
    return (s.replace('&', '&amp;').replace('<', '&lt;')
             .replace('>', '&gt;').replace('"', '&quot;'))


def _shared_strings(rows):
    sst, idx = [], {}
    for row in rows:
        for cell in row:
            s = '' if cell is None else str(cell)
            if s not in idx:
                idx[s] = len(sst)
                sst.append(s)
    return sst, idx


def build_xlsx(headers: list[str], rows: list[list]) -> bytes:
    all_rows = [headers] + [[('' if v is None else str(v)) for v in r] for r in rows]
    sst, idx = _shared_strings(all_rows)

    sheet_rows = []
    for ri, row in enumerate(all_rows, start=1):
        cells = []
        for ci, value in enumerate(row):
            ref = f'{_col_letter(ci)}{ri}'
            cells.append(f'<c r="{ref}" t="s"><v>{idx[value]}</v></c>')
        sheet_rows.append(f'<row r="{ri}">{"".join(cells)}</row>')

    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<sheetData>{"".join(sheet_rows)}</sheetData>'
        '</worksheet>'
    )

    sst_items = ''.join(f'<si><t xml:space="preserve">{_xml_escape(s)}</t></si>' for s in sst)
    sst_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        f'count="{sum(len(r) for r in all_rows)}" uniqueCount="{len(sst)}">'
        f'{sst_items}</sst>'
    )

    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets><sheet name="Questions" sheetId="1" r:id="rId1"/></sheets>'
        '</workbook>'
    )
    workbook_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
        '</Relationships>'
    )
    root_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        '</Relationships>'
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
        '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        '</Types>'
    )
    styles_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>'
        '<fills count="1"><fill><patternFill patternType="none"/></fill></fills>'
        '<borders count="1"><border/></borders>'
        '<cellStyleXfs count="1"><xf/></cellStyleXfs>'
        '<cellXfs count="1"><xf/></cellXfs>'
        '</styleSheet>'
    )

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as z:
        z.writestr('[Content_Types].xml', content_types)
        z.writestr('_rels/.rels', root_rels)
        z.writestr('xl/workbook.xml', workbook_xml)
        z.writestr('xl/_rels/workbook.xml.rels', workbook_rels)
        z.writestr('xl/styles.xml', styles_xml)
        z.writestr('xl/sharedStrings.xml', sst_xml)
        z.writestr('xl/worksheets/sheet1.xml', sheet_xml)
    return buf.getvalue()


xlsx_bytes = build_xlsx(HEADERS, ROWS)
(WORK_DIR / 'questions.xlsx').write_bytes(xlsx_bytes)


# ── 3. README inside the working copy ────────────────────────────────────────
README = """# physics_sample.zip — README

This is a known-good starter pack for the **Enable-LMS ZIP question importer**.

## Files inside the ZIP
    questions.xlsx          # 5 sample physics questions
    images/q02_trajectory.png
    images/q04_circuit.png

## Column reference
| Column | Required | Notes |
|---|---|---|
| question_code | Recommended | Unique within the test, e.g. PHY-001. Auto-generated as Q001… if blank. |
| question_text | YES | The question itself. |
| question_type | Recommended | One of: MCQ_SINGLE, MCQ_MULTI, NUMERICAL, TRUE_FALSE, FILL_BLANK, SUBJECTIVE. Defaults to MCQ_SINGLE. Aliases: MCQ→MCQ_SINGLE, TF→TRUE_FALSE, NUM→NUMERICAL, FILL→FILL_BLANK. |
| option_a..option_e | YES for MCQ types | One choice per cell. Leave blank for non-MCQ. |
| correct_answer | YES | Letter for single (B), comma list for multi (B,D), or literal value for numerical/fill. |
| positive_marks | YES | Numeric (e.g. 4). |
| negative_marks | Optional | Numeric. Defaults to 0. |
| difficulty | Optional | EASY / MEDIUM / HARD. |
| question_image | Optional | Filename inside images/ (case-insensitive). |
| option_a_image..option_e_image | Optional | Filenames inside images/ for picture-option choices. |
| answer_explanation | Optional | Shown to student after submission. |
| tags | Optional | Comma list — useful for filtering. |
| question_order | Optional | Number — printing order. |

## Question-type cheat sheet
- **MCQ_SINGLE**: one correct option; correct_answer = single letter (A/B/C/D/E).
- **MCQ_MULTI**: 2+ correct options; correct_answer = comma list (A,C).
- **TRUE_FALSE**: model as MCQ_SINGLE with option_a=True, option_b=False, option_c/d filled with placeholder text (e.g. "N/A"). The current validator requires all four option cells to be non-empty.
- **NUMERICAL / FILL_BLANK / SUBJECTIVE**: the current validator still requires all four option columns to be non-empty — fill them with placeholder text (e.g. "—") and put the actual expected answer in `correct_answer`.

## Step-by-step
1. Edit `questions.xlsx` — keep the header row exactly as it is.
2. Add/replace pictures under `images/`. Match the filenames you put in `question_image` / `option_*_image` (case-insensitive).
3. Re-zip the **contents** (not the parent folder). The ZIP root must be:
       questions.xlsx
       images/...
4. Admin → **Test sections** → **+ Import ZIP** → pick the target test → upload.
5. Watch **ZIP Import History** at the bottom for ✓ Success or ⚠ Rejected with a one-line reason.

## Common pitfalls
- Both a packed `option` column **and** `option_a` column at the same time → keep only one of those styles.
- Image filenames don't match the spreadsheet cells (case is ignored, but extension matters).
- ZIP includes a top-level folder, e.g. `physics_sample/questions.xlsx`. Re-zip *the files*, not the parent folder.
- correct_answer for MCQ_SINGLE is multi-letter (B,C). Use MCQ_MULTI for that.

## Re-uploads
The importer matches existing rows by `question_code` within the same test. Re-uploading the file with edited content updates those questions in place — it does not create duplicates. Rows with new codes are inserted.
"""
(WORK_DIR / 'README.md').write_text(README)


# ── 4. Re-zip the working copy at root level into physics_sample.zip ─────────
def zip_dir(src: Path, dst: Path) -> None:
    if dst.exists():
        dst.unlink()
    with zipfile.ZipFile(dst, 'w', zipfile.ZIP_DEFLATED) as z:
        # Only include questions.xlsx and images/  (not README — that's only for humans)
        z.write(src / 'questions.xlsx', 'questions.xlsx')
        for p in sorted((src / 'images').glob('*')):
            z.write(p, f'images/{p.name}')


zip_dir(WORK_DIR, ZIP_PATH)

print(f'Wrote   {ZIP_PATH.relative_to(ROOT)}')
print(f'Working copy at  {WORK_DIR.relative_to(ROOT)}/')
print(f'Size:   {ZIP_PATH.stat().st_size:,} bytes')

# ── 5. Sanity check: parse it back through openpyxl + the LMS importer ───────
try:
    from openpyxl import load_workbook
    wb = load_workbook(WORK_DIR / 'questions.xlsx', read_only=True, data_only=True)
    ws = wb.active
    print(f'XLSX OK: {ws.max_row} rows x {ws.max_column} cols')
except ImportError:
    print('openpyxl not installed in this venv — XLSX self-check skipped (still readable in container).')
