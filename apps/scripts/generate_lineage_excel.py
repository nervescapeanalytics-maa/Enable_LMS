#!/usr/bin/env python
"""
Generate an Excel workbook with complete data lineage:
  Sheet 1: All Models & Tables   – model, db_table, app, field inventory
  Sheet 2: Relationships (FK/M2M) – every foreign-key and many-to-many link
  Sheet 3: Constraints            – unique_together, UniqueConstraint, indexes
  Sheet 4: Lineage Diagram (text) – ASCII/Mermaid ER per app

Run:  cd /u01/app/LMS-Portal/apps && python scripts/generate_lineage_excel.py
"""
import os, sys, django

# Bootstrap Django
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'lms_enterprise.settings.base')
django.setup()

from django.apps import apps
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Apps to include (in sidebar order) ──────────────────────────
APP_ORDER = [
    'academics', 'accounts', 'assessments', 'attendance',
    'classes', 'communication', 'materials', 'realtime',
    'scheduling', 'sessions_tracking', 'system_config',
    'tenants', 'eduassist', 'audit',
]

# ── Styling ─────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="844FC1", end_color="844FC1", fill_type="solid")
SECTION_FILL = PatternFill(start_color="E8DEF8", end_color="E8DEF8", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11, color="4A148C")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def style_cell(ws, row, col, value, align=ALIGN_LEFT):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = align
    cell.border = THIN_BORDER
    return cell


def auto_width(ws, max_w=55):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[letter].width = min(length + 3, max_w)


def field_type_str(field):
    """Return a human-readable type string for a Django model field."""
    name = type(field).__name__
    extras = []
    if hasattr(field, 'max_length') and field.max_length:
        extras.append(str(field.max_length))
    if hasattr(field, 'max_digits') and field.max_digits:
        extras.append(f"{field.max_digits},{field.decimal_places}")
    suffix = f"({','.join(extras)})" if extras else ''
    return f"{name}{suffix}"


def get_models_for_app(app_label):
    try:
        cfg = apps.get_app_config(app_label)
        return list(cfg.get_models())
    except LookupError:
        return []


# ════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ════════════════════════════════════════════════════════════════
wb = Workbook()

# ── Sheet 1: Models & Fields ───────────────────────────────────
ws1 = wb.active
ws1.title = "Models & Fields"
headers1 = ["App", "Model", "DB Table", "Field", "Type", "PK", "Nullable",
            "Default", "Unique", "Description / Choices"]
for c, h in enumerate(headers1, 1):
    ws1.cell(row=1, column=c, value=h)
style_header(ws1, 1, len(headers1))

row1 = 2
for app_label in APP_ORDER:
    models = get_models_for_app(app_label)
    if not models:
        continue
    for model in sorted(models, key=lambda m: m.__name__):
        meta = model._meta
        # Section separator
        cell = ws1.cell(row=row1, column=1, value=app_label)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell = ws1.cell(row=row1, column=2, value=model.__name__)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell = ws1.cell(row=row1, column=3, value=meta.db_table)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        for c in range(4, len(headers1) + 1):
            ws1.cell(row=row1, column=c).fill = SECTION_FILL
        for c in range(1, len(headers1) + 1):
            ws1.cell(row=row1, column=c).border = THIN_BORDER
        row1 += 1

        for field in meta.get_fields():
            if field.is_relation and not field.concrete:
                continue  # skip reverse relations
            fname = field.name
            ftype = field_type_str(field) if hasattr(field, 'get_internal_type') else 'Relation'
            pk = 'Yes' if getattr(field, 'primary_key', False) else ''
            nullable = 'Yes' if getattr(field, 'null', False) else ''
            default = ''
            if hasattr(field, 'default') and field.default is not None:
                d = field.default
                if callable(d):
                    default = d.__name__ if hasattr(d, '__name__') else str(d)
                elif d != '':
                    default = str(d)
            unique = 'Yes' if getattr(field, 'unique', False) else ''
            desc = ''
            if hasattr(field, 'choices') and field.choices:
                try:
                    choices_list = list(field.choices)
                    desc = ', '.join(str(c[0]) for c in choices_list[:10])
                    if len(choices_list) > 10:
                        desc += '...'
                except Exception:
                    desc = 'choices'
            if hasattr(field, 'related_model') and field.related_model:
                ftype = f"FK → {field.related_model.__name__}"
                on_del = getattr(field, 'remote_field', None)
                if on_del and hasattr(on_del, 'on_delete') and on_del.on_delete:
                    desc = f"on_delete={on_del.on_delete.__name__}"
                else:
                    desc = ''

            style_cell(ws1, row1, 1, app_label)
            style_cell(ws1, row1, 2, model.__name__)
            style_cell(ws1, row1, 3, meta.db_table)
            style_cell(ws1, row1, 4, fname)
            style_cell(ws1, row1, 5, ftype)
            style_cell(ws1, row1, 6, pk, ALIGN_CENTER)
            style_cell(ws1, row1, 7, nullable, ALIGN_CENTER)
            style_cell(ws1, row1, 8, default)
            style_cell(ws1, row1, 9, unique, ALIGN_CENTER)
            style_cell(ws1, row1, 10, desc)
            row1 += 1

ws1.auto_filter.ref = f"A1:J{row1 - 1}"
auto_width(ws1)

# ── Sheet 2: Relationships ─────────────────────────────────────
ws2 = wb.create_sheet("Relationships")
headers2 = ["#", "Source App", "Source Model", "Source Table", "Field",
            "Relation Type", "Target Model", "Target Table", "on_delete", "Notes"]
for c, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=c, value=h)
style_header(ws2, 1, len(headers2))

row2 = 2
rel_num = 0
for app_label in APP_ORDER:
    models = get_models_for_app(app_label)
    for model in sorted(models, key=lambda m: m.__name__):
        meta = model._meta
        for field in meta.get_fields():
            if not field.is_relation or not field.concrete:
                continue
            rel_num += 1
            target = field.related_model
            if target is None:
                continue
            on_del = ''
            rel_type = type(field).__name__
            if hasattr(field, 'remote_field') and field.remote_field:
                rf = field.remote_field
                if hasattr(rf, 'on_delete') and rf.on_delete:
                    on_del = rf.on_delete.__name__
            if hasattr(field, 'many_to_many') and field.many_to_many:
                rel_type = 'ManyToManyField'
                through = getattr(field.remote_field, 'through', None)
                notes = f"through={through.__name__}" if through and through._meta.auto_created is False else ''
            else:
                notes = ''

            style_cell(ws2, row2, 1, rel_num, ALIGN_CENTER)
            style_cell(ws2, row2, 2, app_label)
            style_cell(ws2, row2, 3, model.__name__)
            style_cell(ws2, row2, 4, meta.db_table)
            style_cell(ws2, row2, 5, field.name)
            style_cell(ws2, row2, 6, rel_type)
            style_cell(ws2, row2, 7, target.__name__)
            style_cell(ws2, row2, 8, target._meta.db_table)
            style_cell(ws2, row2, 9, on_del)
            style_cell(ws2, row2, 10, notes)
            row2 += 1

ws2.auto_filter.ref = f"A1:J{row2 - 1}"
auto_width(ws2)

# ── Sheet 3: Constraints ──────────────────────────────────────
ws3 = wb.create_sheet("Constraints")
headers3 = ["App", "Model", "DB Table", "Constraint Type", "Name", "Fields / Expression"]
for c, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=c, value=h)
style_header(ws3, 1, len(headers3))

row3 = 2
for app_label in APP_ORDER:
    models = get_models_for_app(app_label)
    for model in sorted(models, key=lambda m: m.__name__):
        meta = model._meta
        # unique_together
        if meta.unique_together:
            for ut in meta.unique_together:
                style_cell(ws3, row3, 1, app_label)
                style_cell(ws3, row3, 2, model.__name__)
                style_cell(ws3, row3, 3, meta.db_table)
                style_cell(ws3, row3, 4, 'unique_together')
                style_cell(ws3, row3, 5, '')
                style_cell(ws3, row3, 6, ', '.join(ut))
                row3 += 1
        # Meta.constraints
        for constr in meta.constraints:
            ctype = type(constr).__name__
            cname = getattr(constr, 'name', '')
            if hasattr(constr, 'fields'):
                cfields = ', '.join(constr.fields)
            elif hasattr(constr, 'check'):
                cfields = str(constr.check)
            else:
                cfields = ''
            style_cell(ws3, row3, 1, app_label)
            style_cell(ws3, row3, 2, model.__name__)
            style_cell(ws3, row3, 3, meta.db_table)
            style_cell(ws3, row3, 4, ctype)
            style_cell(ws3, row3, 5, cname)
            style_cell(ws3, row3, 6, cfields)
            row3 += 1
        # Meta.indexes
        for idx in meta.indexes:
            iname = getattr(idx, 'name', '')
            ifields = ', '.join(idx.fields) if hasattr(idx, 'fields') else ''
            style_cell(ws3, row3, 1, app_label)
            style_cell(ws3, row3, 2, model.__name__)
            style_cell(ws3, row3, 3, meta.db_table)
            style_cell(ws3, row3, 4, 'Index')
            style_cell(ws3, row3, 5, iname)
            style_cell(ws3, row3, 6, ifields)
            row3 += 1

ws3.auto_filter.ref = f"A1:F{row3 - 1}"
auto_width(ws3)

# ── Sheet 4: Lineage / ER Diagram (Mermaid) ───────────────────
ws4 = wb.create_sheet("Lineage Diagram")
ws4.cell(row=1, column=1, value="Mermaid ER Diagram").font = Font(bold=True, size=14)
ws4.cell(row=2, column=1, value="Copy the text below into https://mermaid.live to view the interactive diagram.").font = Font(italic=True, color="666666")

lines = ["erDiagram"]

for app_label in APP_ORDER:
    models = get_models_for_app(app_label)
    if not models:
        continue
    lines.append(f"\n    %% ── {app_label.upper()} ──")
    for model in sorted(models, key=lambda m: m.__name__):
        meta = model._meta
        entity = model.__name__
        # Add entity fields
        fields_lines = []
        for field in meta.get_fields():
            if field.is_relation and not field.concrete:
                continue
            fname = field.name
            ftype = field.get_internal_type() if hasattr(field, 'get_internal_type') else 'relation'
            pk_mark = ' PK' if getattr(field, 'primary_key', False) else ''
            fk_mark = ' FK' if (field.is_relation and field.concrete) else ''
            fields_lines.append(f"        {ftype} {fname}{pk_mark}{fk_mark}")
        lines.append(f"    {entity} {{")
        lines.extend(fields_lines[:20])  # cap to prevent huge diagrams
        if len(fields_lines) > 20:
            lines.append(f"        string __{len(fields_lines)-20}_more_fields__")
        lines.append("    }")

    # Relationships
    for model in sorted(models, key=lambda m: m.__name__):
        for field in model._meta.get_fields():
            if not field.is_relation or not field.concrete:
                continue
            target = field.related_model
            if target is None:
                continue
            src = model.__name__
            tgt = target.__name__
            if hasattr(field, 'many_to_many') and field.many_to_many:
                lines.append(f'    {src} }}o--o{{ {tgt} : "{field.name}"')
            else:
                null = getattr(field, 'null', False)
                card = "|o" if null else "||"
                lines.append(f'    {tgt} {card}--o{{ {src} : "{field.name}"')

mermaid_text = "\n".join(lines)

row4 = 4
for line in mermaid_text.split("\n"):
    ws4.cell(row=row4, column=1, value=line)
    row4 += 1

ws4.column_dimensions['A'].width = 120

# ── Sheet 5: App Summary ──────────────────────────────────────
ws5 = wb.create_sheet("App Summary")
headers5 = ["App Label", "Verbose Name", "Models", "Tables", "Total Fields",
            "FK Relations", "Description"]
for c, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=c, value=h)
style_header(ws5, 1, len(headers5))

APP_DESCRIPTIONS = {
    'academics': 'Academic structure: sessions, batches, subjects, chapters, topics, students (Users), reference data',
    'accounts': 'Identity & access management: Student, Teacher, Admin, Parent, roles, permissions, RBAC, security policies',
    'assessments': 'Tests & exams: questions, test attempts, answers, scoring, proctoring, offline marks',
    'attendance': 'Attendance tracking: daily records, correction requests, monthly summaries',
    'classes': 'Live classes & YouTube: scheduled classes, YouTube channels, access tokens, watch time analytics',
    'communication': 'Messaging: announcements, direct messages, notifications, support tickets',
    'materials': 'Study materials: PDFs, videos, presentations, material access tracking',
    'realtime': 'WebSocket events: real-time event delivery and tracking',
    'scheduling': 'Recurring rules, enrollment, blackout dates, class-level attendance',
    'sessions_tracking': 'User sessions: devices, login history, session management, user activity',
    'system_config': 'System settings: feature flags, MFA policies, integrations, AI features, attendance rules, website config',
    'tenants': 'Multi-tenancy: tenant management, domains, billing, feature toggles',
    'eduassist': 'AI chat assistant: conversations, messages, AI-powered educational support',
    'audit': 'Audit trail: comprehensive logging of data changes and system events',
}

row5 = 2
for app_label in APP_ORDER:
    models = get_models_for_app(app_label)
    if not models:
        continue
    total_fields = 0
    total_fks = 0
    tables = set()
    for m in models:
        tables.add(m._meta.db_table)
        for f in m._meta.get_fields():
            if f.is_relation and not f.concrete:
                continue
            total_fields += 1
            if f.is_relation and f.concrete:
                total_fks += 1

    try:
        verbose = apps.get_app_config(app_label).verbose_name
    except Exception:
        verbose = app_label

    style_cell(ws5, row5, 1, app_label)
    style_cell(ws5, row5, 2, verbose)
    style_cell(ws5, row5, 3, len(models), ALIGN_CENTER)
    style_cell(ws5, row5, 4, len(tables), ALIGN_CENTER)
    style_cell(ws5, row5, 5, total_fields, ALIGN_CENTER)
    style_cell(ws5, row5, 6, total_fks, ALIGN_CENTER)
    style_cell(ws5, row5, 7, APP_DESCRIPTIONS.get(app_label, ''))
    row5 += 1

auto_width(ws5)

# ── Save ───────────────────────────────────────────────────────
outpath = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       'docs', 'LMS_Data_Lineage.xlsx')
os.makedirs(os.path.dirname(outpath), exist_ok=True)
wb.save(outpath)
print(f"\n✅  Lineage workbook saved: {outpath}")
print(f"    Sheets: {[ws.title for ws in wb.worksheets]}")
print(f"    Models & Fields rows:  {row1 - 2}")
print(f"    Relationships:         {rel_num}")
print(f"    Constraints rows:      {row3 - 2}")
