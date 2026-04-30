"""Import / back-fill Gender and Category for existing Student records.

Reads a Students workbook (default sheet 0) and matches rows to existing
Student records by phone number (the most stable identifier in the source
spreadsheet). For each matched record, fills in:

  - `gender`   (mapped to Student.Gender choices)
  - `category` (FK to academics.Category, auto-creating per tenant)

Usage:
    python manage.py import_student_demographics --file /path/to/students.xlsx
    python manage.py import_student_demographics --file /path/to/students.xlsx --dry-run
    python manage.py import_student_demographics --file /path/to/students.xlsx --match-on email

Idempotent: re-running on the same file is safe and only updates rows whose
target columns are NULL/blank (use --overwrite to force).
"""

from __future__ import annotations

import os
from collections import Counter
from typing import Optional

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone


GENDER_MAP = {
    'M': 'MALE', 'MALE': 'MALE',
    'F': 'FEMALE', 'FEMALE': 'FEMALE',
    'O': 'OTHER', 'OTHER': 'OTHER',
    'PNTS': 'PREFER_NOT_TO_SAY',
    'PREFER NOT TO SAY': 'PREFER_NOT_TO_SAY',
    'PREFER_NOT_TO_SAY': 'PREFER_NOT_TO_SAY',
}

# Canonicalize messy category labels found in spreadsheets.
CATEGORY_CANONICAL = {
    'GEN': 'General',
    'GENERAL': 'General',
    'OBC': 'OBC',
    'SC': 'SC',
    'ST': 'ST',
    'EWS': 'EWS',
    'GEN-EWS': 'EWS',
    'GENEWS': 'EWS',
    'GENERAL-EWS': 'EWS',
}

# Values that mean "no real category" — skipped entirely.
CATEGORY_BLANKS = {'', 'NA', 'N/A', 'NONE', 'CHOOSE CATEGORY', '-', '--'}


def _norm_phone(v) -> str:
    """Normalize a phone string to compare against Student.phone."""
    if v is None:
        return ''
    s = str(v).strip()
    if s.endswith('.0'):  # excel-imported floats
        s = s[:-2]
    # Keep only digits — many spreadsheets prepend '+91' or have separators
    digits = ''.join(c for c in s if c.isdigit())
    # Last 10 digits is the canonical Indian phone number
    return digits[-10:] if len(digits) >= 10 else digits


def _norm_gender(v) -> Optional[str]:
    if v is None:
        return None
    key = str(v).strip().upper()
    return GENDER_MAP.get(key)


def _norm_category(v) -> Optional[str]:
    if v is None:
        return None
    key = str(v).strip()
    if not key:
        return None
    upper = key.upper()
    if upper in CATEGORY_BLANKS:
        return None
    return CATEGORY_CANONICAL.get(upper, key.title())


class Command(BaseCommand):
    help = 'Back-fill Gender and Category on Student records from an Excel sheet.'

    def add_arguments(self, parser):
        parser.add_argument('--file', required=True, help='Path to .xlsx file')
        parser.add_argument('--sheet', default=None, help='Sheet name (defaults to first)')
        parser.add_argument('--match-on', default='phone', choices=['phone', 'email'],
                            help='Column used to match spreadsheet rows to Student records')
        parser.add_argument('--overwrite', action='store_true',
                            help='Overwrite gender/category even if already set')
        parser.add_argument('--dry-run', action='store_true',
                            help='Read + report only; no DB writes')
        parser.add_argument('--tenant', default=None,
                            help='Tenant code — required for creating new Category rows. '
                                 'Defaults to the Student\'s own tenant.')

    def handle(self, *args, **opts):
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise CommandError('openpyxl is required: pip install openpyxl') from e

        path = opts['file']
        if not os.path.exists(path):
            raise CommandError(f'File not found: {path}')

        from accounts.models import Student
        from academics.models import Category

        wb = load_workbook(filename=path, read_only=True, data_only=True)
        ws = wb[opts['sheet']] if opts['sheet'] else wb[wb.sheetnames[0]]
        self.stdout.write(self.style.NOTICE(
            f'Reading sheet "{ws.title}" from {path}'
        ))

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows_iter))
        except StopIteration:
            raise CommandError('Empty sheet')

        def find_col(*names):
            lc = [str(h).strip().lower() if h else '' for h in header]
            for nm in names:
                if nm in lc:
                    return lc.index(nm)
            return None

        ix_phone = find_col('phone', 'mobile', 'mobile number', 'guardian contact no')
        ix_email = find_col('email')
        ix_gender = find_col('gender')
        ix_category = find_col('category')
        ix_district = find_col('district', 'district name')
        ix_city = find_col('city village', 'city/village', 'city village ', 'city_village', 'city')
        ix_marks = find_col('previous class marks', 'previous_class_marks', 'prev class marks')
        ix_match = ix_phone if opts['match_on'] == 'phone' else ix_email

        if ix_match is None:
            raise CommandError(f'Match column "{opts["match_on"]}" not found in header: {header}')
        if ix_gender is None and ix_category is None and ix_district is None and ix_marks is None:
            raise CommandError('No importable column (gender/category/district/previous class marks) found.')

        # Build a category cache per (tenant_id, canonical_name).
        tenant_override = None
        if opts['tenant']:
            from tenants.models import Tenant
            tenant_override = Tenant.objects.get(code=opts['tenant'])

        cat_cache: dict[tuple, Category] = {}

        def get_or_create_category(tenant, name: str) -> Category:
            key = (str(tenant.id) if tenant else None, name)
            if key in cat_cache:
                return cat_cache[key]
            obj, _ = Category.objects.get_or_create(
                tenant=tenant, name=name,
                defaults={'status': 'Active', 'show_in_student': True},
            )
            cat_cache[key] = obj
            return obj

        # Stats
        seen = 0
        matched = 0
        no_match = 0
        gender_set = 0
        category_set = 0
        district_set = 0
        city_set = 0
        marks_set = 0
        skipped_filled = 0
        bad_gender = Counter()
        bad_category = Counter()
        match_misses_sample = []

        # Pre-load all phone→student in chunks (41k rows is small)
        match_field = 'phone' if opts['match_on'] == 'phone' else 'email'
        students_by_key: dict[str, list] = {}
        for s in Student.objects.all().only(
            'id', 'phone', 'email', 'tenant_id', 'gender', 'category_id',
            'district', 'city', 'previous_class_marks',
        ):
            raw = getattr(s, match_field) or ''
            key = _norm_phone(raw) if match_field == 'phone' else str(raw).strip().lower()
            if not key:
                continue
            students_by_key.setdefault(key, []).append(s)

        self.stdout.write(self.style.NOTICE(
            f'Indexed {sum(len(v) for v in students_by_key.values())} students by {match_field} '
            f'({len(students_by_key)} unique keys)'
        ))

        # ── Walk the sheet ──
        to_update: dict[str, Student] = {}
        total_written = 0
        FIELDS = ['gender', 'category', 'district', 'city', 'previous_class_marks']

        def _flush_chunk():
            nonlocal total_written, to_update
            if not to_update or opts['dry_run']:
                return
            with transaction.atomic():
                Student.objects.bulk_update(list(to_update.values()), FIELDS, batch_size=500)
            total_written += len(to_update)
            to_update = {}

        for row_num, row in enumerate(rows_iter, start=2):
            seen += 1
            if not row:
                continue

            raw_match = row[ix_match] if ix_match < len(row) else None
            mkey = _norm_phone(raw_match) if match_field == 'phone' else str(raw_match or '').strip().lower()
            if not mkey:
                no_match += 1
                continue

            candidates = students_by_key.get(mkey)
            if not candidates:
                no_match += 1
                if len(match_misses_sample) < 5:
                    match_misses_sample.append((row_num, raw_match))
                continue

            new_gender = None
            if ix_gender is not None and ix_gender < len(row):
                raw_g = row[ix_gender]
                if raw_g not in (None, ''):
                    new_gender = _norm_gender(raw_g)
                    if not new_gender:
                        bad_gender[str(raw_g).strip().upper()] += 1

            new_category_name = None
            if ix_category is not None and ix_category < len(row):
                raw_c = row[ix_category]
                if raw_c not in (None, ''):
                    new_category_name = _norm_category(raw_c)
                    if not new_category_name and str(raw_c).strip().upper() not in CATEGORY_BLANKS:
                        bad_category[str(raw_c).strip().upper()] += 1

            new_district = None
            if ix_district is not None and ix_district < len(row):
                raw_d = row[ix_district]
                if raw_d not in (None, ''):
                    d = str(raw_d).strip()
                    if d:
                        new_district = d[:100]

            new_city = None
            if ix_city is not None and ix_city < len(row):
                raw_cv = row[ix_city]
                if raw_cv not in (None, ''):
                    cv = str(raw_cv).strip()
                    if cv:
                        new_city = cv[:100]

            new_marks = None
            if ix_marks is not None and ix_marks < len(row):
                raw_m = row[ix_marks]
                if raw_m not in (None, ''):
                    m = str(raw_m).strip()
                    if m:
                        new_marks = m[:100]

            for s in candidates:
                matched += 1
                changed = False

                if new_gender and (opts['overwrite'] or not s.gender):
                    s.gender = new_gender
                    gender_set += 1
                    changed = True
                elif new_gender and s.gender:
                    skipped_filled += 1

                if new_category_name and (opts['overwrite'] or not s.category_id):
                    tenant = tenant_override if tenant_override else getattr(s, 'tenant', None)
                    if tenant is None:
                        # Cannot create a tenant-scoped Category without a tenant
                        bad_category[f'NO_TENANT:{new_category_name}'] += 1
                    else:
                        cat = get_or_create_category(tenant, new_category_name)
                        s.category = cat
                        category_set += 1
                        changed = True
                elif new_category_name and s.category_id:
                    skipped_filled += 1

                if new_district and (opts['overwrite'] or not s.district):
                    s.district = new_district
                    district_set += 1
                    changed = True
                elif new_district and s.district:
                    skipped_filled += 1

                # City: ALWAYS overwrite — the original loader incorrectly
                # populated `city` with the District value, so we must fix it.
                if new_city and s.city != new_city:
                    s.city = new_city
                    city_set += 1
                    changed = True

                if new_marks and (opts['overwrite'] or not s.previous_class_marks):
                    s.previous_class_marks = new_marks
                    marks_set += 1
                    changed = True
                elif new_marks and s.previous_class_marks:
                    skipped_filled += 1

                if changed:
                    to_update[str(s.id)] = s
                    if len(to_update) >= 2000:
                        _flush_chunk()

        # Final flush
        _flush_chunk()

        # ── Report ──
        self.stdout.write(self.style.NOTICE(
            f'\nRows in sheet (excluding header): {seen}'
        ))
        self.stdout.write(f'  matched           : {matched}')
        self.stdout.write(f'  no-match          : {no_match}')
        self.stdout.write(f'  gender to set     : {gender_set}')
        self.stdout.write(f'  category to set   : {category_set}')
        self.stdout.write(f'  district to set   : {district_set}')
        self.stdout.write(f'  city to set       : {city_set}')
        self.stdout.write(f'  prev marks to set : {marks_set}')
        self.stdout.write(f'  skipped (filled)  : {skipped_filled}')
        if bad_gender:
            self.stdout.write(self.style.WARNING(
                f'  unrecognized gender values: {dict(bad_gender)}'
            ))
        if bad_category:
            self.stdout.write(self.style.WARNING(
                f'  unrecognized category values: {dict(bad_category)}'
            ))
        if match_misses_sample:
            self.stdout.write(self.style.WARNING(
                f'  sample no-match keys: {match_misses_sample}'
            ))

        if opts['dry_run']:
            self.stdout.write(self.style.NOTICE('\n--dry-run: no DB writes performed.'))
            return

        self.stdout.write(self.style.SUCCESS(
            f'\nUpdated {total_written} student records '
            f'(gender:{gender_set}, category:{category_set}, '
            f'district:{district_set}, city:{city_set}, '
            f'prev_marks:{marks_set}) '
            f'at {timezone.now():%Y-%m-%d %H:%M:%S}'
        ))
