"""
Question bulk import — CSV / XLSX / ZIP.

Public API:
    parse_rows(file, ext)             -> list[dict]
    validate_rows(rows, *, test, tenant) -> (cleaned, errors)
    apply_rows(rows, *, test, tenant, actor=None, request=None) -> dict

The "dry-run preview" UI feeds the user through ``parse_rows`` then
``validate_rows`` without ever calling ``apply_rows``.

Collision policy: **update on collision** (matched by ``question_code``
within the same ``test``). Wrapped in a single transaction.

ZIP layout (expected):
    questions.xlsx        (or questions.csv)   — required
    images/               (optional)           — referenced by columns
                                                  ``question_image``,
                                                  ``option_a_image`` …
                                                  ``option_e_image``
                                                  using filenames like
                                                  ``q1.png`` (relative).
"""
from __future__ import annotations

import csv
import io
import logging
import os
import shutil
import tempfile
import uuid
import zipfile
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable, Optional

from django.conf import settings
from django.db import transaction

from .models import Question, Test
from .permissions import log_exam_event

logger = logging.getLogger(__name__)


REQUIRED_COLS = [
    'question_code', 'question_text', 'question_type',
    'option_a', 'option_b', 'option_c', 'option_d',
    'correct_answer', 'positive_marks',
]
OPTIONAL_COLS = [
    'option_e', 'difficulty', 'negative_marks', 'partial_marks',
    'answer_explanation', 'question_order', 'tags',
    'question_image', 'option_a_image', 'option_b_image',
    'option_c_image', 'option_d_image', 'option_e_image',
]
ALL_COLS = REQUIRED_COLS + OPTIONAL_COLS

# Columns whose value (when imported from a ZIP) is a relative filename
# inside the ZIP that we publish as a media URL on the Question record.
IMAGE_COLS = (
    'question_image',
    'option_a_image', 'option_b_image', 'option_c_image',
    'option_d_image', 'option_e_image',
)

VALID_TYPES = {
    'MCQ_SINGLE', 'MCQ_MULTI', 'NUMERICAL', 'TRUE_FALSE',
    'FILL_BLANK', 'SUBJECTIVE', 'MATRIX_MATCH',
    'ASSERTION_REASON', 'COMPREHENSION',
}
VALID_DIFFICULTY = {'EASY', 'MEDIUM', 'HARD'}


# Aliases — map common header variants to canonical column names so that
# spreadsheets created by humans (Title Case, spaces, dashes, abbreviations)
# import without manual editing.
HEADER_ALIASES = {
    # question_code
    'code': 'question_code', 'qcode': 'question_code', 'q_code': 'question_code',
    'q_no': 'question_code', 'qno': 'question_code', 'qid': 'question_code',
    'question_id': 'question_code', 'id': 'question_code',
    # question_text
    'question': 'question_text', 'text': 'question_text', 'stem': 'question_text',
    'q_text': 'question_text', 'body': 'question_text',
    # question_type
    'type': 'question_type', 'qtype': 'question_type', 'q_type': 'question_type',
    # options
    'a': 'option_a', 'b': 'option_b', 'c': 'option_c', 'd': 'option_d', 'e': 'option_e',
    'opt_a': 'option_a', 'opt_b': 'option_b', 'opt_c': 'option_c',
    'opt_d': 'option_d', 'opt_e': 'option_e',
    'choice_a': 'option_a', 'choice_b': 'option_b', 'choice_c': 'option_c',
    'choice_d': 'option_d', 'choice_e': 'option_e',
    # correct_answer
    'answer': 'correct_answer', 'correct': 'correct_answer',
    'correct_option': 'correct_answer', 'right_answer': 'correct_answer',
    'key': 'correct_answer',
    # marks
    'marks': 'positive_marks', 'positive': 'positive_marks',
    'pos_marks': 'positive_marks', 'mark': 'positive_marks',
    'negative': 'negative_marks', 'neg_marks': 'negative_marks',
    'partial': 'partial_marks',
    # explanation
    'explanation': 'answer_explanation', 'solution': 'answer_explanation',
    # difficulty/level
    'level': 'difficulty',
    # ordering
    'order': 'question_order', 'sequence': 'question_order',
    'sl_no': 'question_order', 'serial': 'question_order',
}


def _canon_header(name: str) -> str:
    """Lowercase, strip, and convert spaces / dashes / dots to underscores;
    then resolve through HEADER_ALIASES to a canonical column name."""
    if not name:
        return ''
    s = str(name).strip().lower()
    # collapse any whitespace, dashes, dots, slashes to '_'
    out = []
    for ch in s:
        if ch.isalnum():
            out.append(ch)
        else:
            out.append('_')
    s = ''.join(out)
    while '__' in s:
        s = s.replace('__', '_')
    s = s.strip('_')
    return HEADER_ALIASES.get(s, s)


def _canon_row(raw: dict) -> dict:
    """Return a row with canonicalised header keys (last-wins on collision)."""
    out = {}
    for k, v in (raw or {}).items():
        if not k:
            continue
        ck = _canon_header(k)
        if not ck:
            continue
        out[ck] = v.strip() if isinstance(v, str) else v
    return out


# ---------------------------------------------------------------------------
# Parse
# ---------------------------------------------------------------------------
def _parse_csv(blob: bytes) -> list[dict]:
    text = blob.decode('utf-8-sig', errors='replace')
    reader = csv.DictReader(io.StringIO(text))
    return [dict(r) for r in reader]


def _parse_xlsx(blob: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError('openpyxl not installed; cannot parse XLSX') from e
    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(c or '').strip() for c in next(rows)]
    except StopIteration:
        return []
    out = []
    for row in rows:
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue
        out.append({header[i]: ('' if c is None else str(c).strip())
                    for i, c in enumerate(row) if i < len(header)})
    return out


def parse_rows(file_obj, ext: str) -> list[dict]:
    blob = file_obj.read() if hasattr(file_obj, 'read') else bytes(file_obj)
    ext = (ext or '').lower().lstrip('.')
    if ext == 'csv':
        return _parse_csv(blob)
    if ext in ('xlsx', 'xls'):
        return _parse_xlsx(blob)
    if ext == 'zip':
        rows, _images_dir = _parse_zip(blob, register=False)
        return rows
    raise ValueError(f'Unsupported file extension: {ext!r}')


# ---------------------------------------------------------------------------
# ZIP support
# ---------------------------------------------------------------------------
ALLOWED_IMAGE_EXTS = {'.png', '.jpg', '.jpeg', '.gif', '.webp', '.svg'}


def _parse_zip(blob: bytes, *, register: bool) -> tuple[list[dict], dict]:
    """Extract a ZIP and return (rows, image_url_map).

    The ZIP must contain exactly one of:
        questions.xlsx, questions.csv, *.xlsx, *.csv

    Image references in IMAGE_COLS are resolved to entries within the ZIP
    (case-insensitive, slash-tolerant). When ``register`` is True, image
    payloads are written under MEDIA_ROOT/question_images/<uuid>/<basename>
    and the row column is rewritten to the corresponding /media/... URL.

    On register=False, image columns are returned as-is (filenames).
    """
    rows: list[dict] = []
    image_url_map: dict[str, str] = {}

    try:
        zf = zipfile.ZipFile(io.BytesIO(blob))
    except zipfile.BadZipFile as e:
        raise ValueError(f'Invalid ZIP archive: {e}') from e

    # Find the spreadsheet
    names = [n for n in zf.namelist() if not n.endswith('/')]

    # ── QTI 1.x / 2.x detection ────────────────────────────────────────────
    # QTI Content Packages always carry an imsmanifest.xml at the root.
    # Convert MCQ items to our row schema in-flight.
    is_qti = any(Path(n).name.lower() == 'imsmanifest.xml' for n in names)
    if is_qti:
        try:
            qti_rows = _parse_qti(zf, names)
        except Exception as e:  # noqa: BLE001
            zf.close()
            raise ValueError(f'QTI parsing failed: {e}') from e
        # Allow image co-registration alongside QTI items
        rows = qti_rows
        sheet_name = None
    else:
        sheet_name = None
        for cand in ('questions.xlsx', 'questions.csv'):
            for n in names:
                if Path(n).name.lower() == cand:
                    sheet_name = n
                    break
            if sheet_name:
                break
        if not sheet_name:
            for n in names:
                low = n.lower()
                if low.endswith('.xlsx') or low.endswith('.csv'):
                    sheet_name = n
                    break
        if not sheet_name:
            raise ValueError('ZIP must contain questions.xlsx, questions.csv, or imsmanifest.xml (QTI)')

        sheet_blob = zf.read(sheet_name)
        if sheet_name.lower().endswith('.csv'):
            rows = _parse_csv(sheet_blob)
        else:
            rows = _parse_xlsx(sheet_blob)

    # Build a case-insensitive basename → zip-entry map for image lookup
    by_basename: dict[str, str] = {}
    for n in names:
        if n == sheet_name:
            continue
        ext = Path(n).suffix.lower()
        if ext in ALLOWED_IMAGE_EXTS:
            by_basename[Path(n).name.lower()] = n

    if not register:
        zf.close()
        return rows, image_url_map

    # Persist images on disk under MEDIA_ROOT/question_images/<batch>/
    media_root = Path(getattr(settings, 'MEDIA_ROOT', 'runtime/media'))
    media_url = (getattr(settings, 'MEDIA_URL', '/media/') or '/media/').rstrip('/')
    batch_dir_name = uuid.uuid4().hex[:12]
    target_dir = media_root / 'question_images' / batch_dir_name
    target_dir.mkdir(parents=True, exist_ok=True)

    for row in rows:
        for col in IMAGE_COLS:
            ref = (row.get(col) or '').strip()
            if not ref:
                continue
            # Resolve by basename (also accept full zip paths)
            base = Path(ref).name.lower()
            zip_entry = by_basename.get(base)
            if not zip_entry:
                # If they wrote a full path with slashes, try a strict match
                norm = ref.replace('\\', '/').lstrip('/').lower()
                for n in names:
                    if n.lower() == norm:
                        zip_entry = n
                        break
            if not zip_entry:
                continue  # validate phase will not flag — image is optional

            cached = image_url_map.get(zip_entry)
            if not cached:
                safe_name = Path(zip_entry).name
                # Sanitise filename (avoid path traversal)
                safe_name = safe_name.replace('/', '_').replace('\\', '_')
                out_path = target_dir / safe_name
                with zf.open(zip_entry) as src, open(out_path, 'wb') as dst:
                    shutil.copyfileobj(src, dst)
                cached = f'{media_url}/question_images/{batch_dir_name}/{safe_name}'
                image_url_map[zip_entry] = cached
            row[col] = cached

    zf.close()
    return rows, image_url_map


def parse_zip_with_assets(file_obj) -> list[dict]:
    """Convenience wrapper used by views that want image extraction."""
    blob = file_obj.read() if hasattr(file_obj, 'read') else bytes(file_obj)
    rows, _ = _parse_zip(blob, register=True)
    return rows


# ---------------------------------------------------------------------------
# QTI 1.x / 2.x  (best-effort MCQ_SINGLE / MCQ_MULTI / TRUE_FALSE)
# ---------------------------------------------------------------------------
def _qti_strip(s):
    if s is None:
        return ''
    return ' '.join(str(s).split()).strip()


def _parse_qti(zf: zipfile.ZipFile, names: list[str]) -> list[dict]:
    """Parse a QTI Content Package (1.x or 2.x) into row dicts.

    Heuristic: search every *.xml entry for ``assessmentItem`` (QTI 2.x) or
    ``item``/``response_lid`` (QTI 1.x). Unsupported item types are skipped
    with a warning rather than failing the whole import.
    """
    import re
    import xml.etree.ElementTree as ET

    out: list[dict] = []
    skipped = 0

    def _localname(tag: str) -> str:
        # strip XML namespaces: '{ns}tag' → 'tag'
        return tag.rsplit('}', 1)[-1].lower()

    for n in names:
        if not n.lower().endswith('.xml'):
            continue
        if Path(n).name.lower() == 'imsmanifest.xml':
            continue
        try:
            data = zf.read(n)
            root = ET.fromstring(data)
        except Exception:  # noqa: BLE001
            continue

        # Walk every node and snapshot QTI items
        for node in root.iter():
            tag = _localname(node.tag)

            # ─── QTI 2.x ────────────────────────────────────────────────
            if tag == 'assessmentitem':
                code = node.attrib.get('identifier') or f'QTI{len(out)+1:03d}'
                # itemBody → text + choiceInteraction
                body_text = ''
                choices: list[tuple[str, str]] = []  # [(identifier, text)]
                correct_ids: list[str] = []
                for sub in node.iter():
                    name = _localname(sub.tag)
                    if name == 'itembody':
                        body_text = _qti_strip(''.join(sub.itertext()))
                    elif name == 'simplechoice':
                        choices.append((
                            sub.attrib.get('identifier', ''),
                            _qti_strip(''.join(sub.itertext())),
                        ))
                    elif name == 'correctresponse':
                        for v in sub.iter():
                            if _localname(v.tag) == 'value':
                                txt = (v.text or '').strip()
                                if txt:
                                    correct_ids.append(txt)
                if not choices:
                    skipped += 1
                    continue
                row = _qti_row_from_choices(code, body_text, choices, correct_ids)
                if row:
                    out.append(row)
                else:
                    skipped += 1

            # ─── QTI 1.x ────────────────────────────────────────────────
            elif tag == 'item':
                code = node.attrib.get('ident') or node.attrib.get('title') or f'QTI{len(out)+1:03d}'
                body_text = ''
                choices: list[tuple[str, str]] = []
                correct_ids: list[str] = []
                for sub in node.iter():
                    name = _localname(sub.tag)
                    if name == 'mattext' and not body_text:
                        body_text = _qti_strip(sub.text or '')
                    elif name == 'response_label':
                        cid = sub.attrib.get('ident', '')
                        text = ''
                        for tt in sub.iter():
                            if _localname(tt.tag) == 'mattext':
                                text = _qti_strip(tt.text or '')
                                break
                        choices.append((cid, text))
                    elif name == 'varequal':
                        txt = (sub.text or '').strip()
                        if txt:
                            correct_ids.append(txt)
                if not choices:
                    skipped += 1
                    continue
                row = _qti_row_from_choices(code, body_text, choices, correct_ids)
                if row:
                    out.append(row)
                else:
                    skipped += 1

    if skipped:
        logger.info('QTI import: skipped %d unsupported item(s)', skipped)
    if not out:
        raise ValueError(
            'QTI package contained no MCQ-style items we could parse.'
        )
    return out


def _qti_row_from_choices(code, body, choices, correct_ids) -> Optional[dict]:
    """Map up to 5 QTI choices → option_a..option_e and resolve correct letters."""
    if not choices:
        return None
    letters = ['A', 'B', 'C', 'D', 'E']
    if len(choices) > 5:
        choices = choices[:5]
    id_to_letter = {cid: letters[i] for i, (cid, _t) in enumerate(choices)}
    correct_letters = sorted({id_to_letter[c] for c in correct_ids if c in id_to_letter})
    if not correct_letters:
        return None
    is_multi = len(correct_letters) > 1
    row = {
        'question_code': str(code)[:50],
        'question_text': body or 'QTI question',
        'question_type': 'MCQ_MULTI' if is_multi else 'MCQ_SINGLE',
        'correct_answer': ','.join(correct_letters),
        'positive_marks': '1',
        'negative_marks': '0',
        'partial_marks': '0',
        'difficulty': 'MEDIUM',
        'question_order': '0',
        'tags': 'qti',
    }
    for i, (_cid, text) in enumerate(choices):
        row[f'option_{letters[i].lower()}'] = text
    for i in range(len(choices), 4):
        row[f'option_{letters[i].lower()}'] = ''
    return row


# ---------------------------------------------------------------------------
# Export — bundle a Test's questions back to a ZIP package
# ---------------------------------------------------------------------------
def export_test_zip(test: Test) -> bytes:
    """Return ZIP bytes containing questions.csv + (referenced) images.

    Image columns that point at /media/... are resolved to filesystem paths
    under MEDIA_ROOT and added to images/ in the ZIP, with the column
    rewritten to a relative path.
    """
    from django.conf import settings as _s

    media_root = Path(getattr(_s, 'MEDIA_ROOT', 'runtime/media'))
    media_url = (getattr(_s, 'MEDIA_URL', '/media/') or '/media/').rstrip('/')

    qs = (Question.objects
          .filter(tenant=test.tenant, test=test, is_deleted=False)
          .order_by('question_order', 'question_code'))

    sio = io.StringIO()
    writer = csv.DictWriter(sio, fieldnames=ALL_COLS)
    writer.writeheader()

    image_map: dict[str, str] = {}  # source_fs_path → arcname
    for q in qs:
        row = {c: '' for c in ALL_COLS}
        row.update({
            'question_code': q.question_code or f'Q{q.question_order or 0}',
            'question_text': q.question_text or '',
            'question_type': q.question_type or 'MCQ_SINGLE',
            'option_a': q.option_a or '', 'option_b': q.option_b or '',
            'option_c': q.option_c or '', 'option_d': q.option_d or '',
            'option_e': q.option_e or '',
            'correct_answer': q.correct_answer or '',
            'positive_marks': str(q.positive_marks or ''),
            'negative_marks': str(q.negative_marks or ''),
            'partial_marks': str(q.partial_marks or ''),
            'difficulty': q.difficulty or '',
            'answer_explanation': q.answer_explanation or '',
            'question_order': str(q.question_order or ''),
            'tags': ','.join(q.tags or []) if isinstance(q.tags, list) else (q.tags or ''),
        })
        for col in IMAGE_COLS:
            url = getattr(q, col, '') or ''
            if not url:
                continue
            arc = None
            if url.startswith(media_url + '/'):
                rel = url[len(media_url) + 1:]
                src = media_root / rel
                if src.exists():
                    arc = f'images/{Path(rel).name}'
                    image_map[str(src)] = arc
                    row[col] = arc
            else:
                row[col] = url  # external URL — leave as-is
        writer.writerow(row)

    zbuf = io.BytesIO()
    # Prepend UTF-8 BOM so Excel auto-detects encoding (avoids mojibake of
    # math symbols like −, π, σ, etc. when opened on Windows).
    csv_bytes = ('\ufeff' + sio.getvalue()).encode('utf-8')
    with zipfile.ZipFile(zbuf, 'w', zipfile.ZIP_DEFLATED) as z:
        z.writestr('questions.csv', csv_bytes)
        for src, arc in image_map.items():
            try:
                z.write(src, arc)
            except OSError:
                continue
        # Bundle a small README for the human admin
        z.writestr('README.txt',
                   'Enable-LMS Test export\n\n'
                   f'Test code: {test.test_code}\nTitle: {test.title}\n'
                   f'Questions: {qs.count()}\n\n'
                   'Edit questions.csv and re-import via /staff/exams/import/.\n'
                   'Image filenames must remain in the images/ folder.\n')
    return zbuf.getvalue()


# ---------------------------------------------------------------------------
# Validate
# ---------------------------------------------------------------------------
def _to_decimal(v, default='0'):
    try:
        return Decimal(str(v).strip() or default)
    except (InvalidOperation, AttributeError):
        return None


def validate_rows(rows: list[dict], *, test: Test, tenant) -> tuple[list[dict], list[dict]]:
    """
    Return (cleaned_rows, errors).
    - cleaned_rows: dicts ready for ``apply_rows``, with an ``_action`` key
      ('create' or 'update') and ``_question_code`` (normalised).
    - errors: list of {row, code, message}.
    """
    cleaned: list[dict] = []
    errors: list[dict] = []

    existing_codes = set(
        Question.objects.filter(test=test, is_deleted=False)
        .values_list('question_code', flat=True)
    )
    seen: set[str] = set()

    for i, raw in enumerate(rows, start=1):
        row = _canon_row(raw)

        # default question_type to MCQ_SINGLE if absent — the most common case
        if not row.get('question_type'):
            row['question_type'] = 'MCQ_SINGLE'

        # required columns present
        missing = [c for c in REQUIRED_COLS if not row.get(c)]
        if missing:
            present = sorted(k for k in row.keys() if not k.startswith('_'))
            errors.append({'row': i, 'code': 'MISSING',
                           'message': (
                               f'missing: {", ".join(missing)}. '
                               f'found columns: {", ".join(present) or "(none)"}'
                           )})
            continue

        qcode = row['question_code'].strip()
        if qcode in seen:
            errors.append({'row': i, 'code': 'DUPLICATE_IN_FILE',
                           'message': f'question_code "{qcode}" repeated in file'})
            continue
        seen.add(qcode)

        qtype = row['question_type'].upper()
        if qtype not in VALID_TYPES:
            errors.append({'row': i, 'code': 'BAD_TYPE',
                           'message': f'question_type "{qtype}" invalid'})
            continue

        diff = (row.get('difficulty') or 'MEDIUM').upper()
        if diff not in VALID_DIFFICULTY:
            errors.append({'row': i, 'code': 'BAD_DIFFICULTY',
                           'message': f'difficulty "{diff}" invalid'})
            continue

        ans = row['correct_answer'].strip().upper()
        if qtype == 'MCQ_SINGLE':
            opts = {k.upper().split('_')[1]: row.get(k)
                    for k in ('option_a', 'option_b', 'option_c', 'option_d', 'option_e')
                    if row.get(k)}
            if ans not in opts:
                errors.append({'row': i, 'code': 'ANSWER_NOT_IN_OPTIONS',
                               'message': f'correct_answer "{ans}" not in provided options'})
                continue

        pos = _to_decimal(row.get('positive_marks'))
        if pos is None:
            errors.append({'row': i, 'code': 'BAD_MARKS',
                           'message': 'positive_marks must be numeric'})
            continue
        neg = _to_decimal(row.get('negative_marks') or '0') or Decimal('0')
        partial = _to_decimal(row.get('partial_marks') or '0') or Decimal('0')

        order_raw = row.get('question_order') or '0'
        try:
            order = int(float(order_raw))
        except (TypeError, ValueError):
            order = 0

        cleaned.append({
            '_question_code': qcode,
            '_action': 'update' if qcode in existing_codes else 'create',
            'question_code': qcode,
            'question_text': row['question_text'],
            'question_type': qtype,
            'difficulty': diff,
            'option_a': row.get('option_a') or None,
            'option_b': row.get('option_b') or None,
            'option_c': row.get('option_c') or None,
            'option_d': row.get('option_d') or None,
            'option_e': row.get('option_e') or None,
            'correct_answer': ans,
            'positive_marks': pos,
            'negative_marks': neg,
            'partial_marks': partial,
            'answer_explanation': row.get('answer_explanation') or '',
            'question_order': order,
            # Image URLs (already resolved to /media/... by parse_zip_with_assets,
            # or empty when CSV/XLSX direct)
            'question_image': row.get('question_image') or None,
            'option_a_image': row.get('option_a_image') or None,
            'option_b_image': row.get('option_b_image') or None,
            'option_c_image': row.get('option_c_image') or None,
            'option_d_image': row.get('option_d_image') or None,
            'option_e_image': row.get('option_e_image') or None,
        })
    return cleaned, errors


# ---------------------------------------------------------------------------
# Apply
# ---------------------------------------------------------------------------
@transaction.atomic
def apply_rows(rows: list[dict], *, test: Test, tenant, actor=None,
               request=None) -> dict:
    created = updated = 0
    for r in rows:
        qcode = r.pop('_question_code')
        action = r.pop('_action')
        defaults = {'tenant': tenant, 'test': test, **r}
        obj, was_created = Question.objects.update_or_create(
            test=test, question_code=qcode, defaults=defaults,
        )
        if was_created:
            created += 1
        else:
            updated += 1

    # Recount
    test.total_questions = Question.objects.filter(test=test, is_deleted=False).count()
    test.total_marks = sum(
        (q.positive_marks or 0)
        for q in Question.objects.filter(test=test, is_deleted=False)
    )
    test.save(update_fields=['total_questions', 'total_marks'])

    log_exam_event(
        request=request, actor=actor,
        action='QUESTION_BULK_IMPORT',
        resource_type='Test', resource_id=test.id, resource_name=test.test_code,
        description=f'Imported {created + updated} rows ({created} new, {updated} updated)',
        extra_meta={'created': created, 'updated': updated},
    )
    return {'created': created, 'updated': updated, 'total': created + updated}


# ---------------------------------------------------------------------------
# Template (for download)
# ---------------------------------------------------------------------------
def template_csv() -> str:
    sample = {
        'question_code': 'IMP-001',
        'question_text': 'Capital of France?',
        'question_type': 'MCQ_SINGLE',
        'difficulty': 'EASY',
        'option_a': 'Berlin', 'option_b': 'Madrid',
        'option_c': 'Paris', 'option_d': 'Rome', 'option_e': '',
        'correct_answer': 'C',
        'positive_marks': '4', 'negative_marks': '-1', 'partial_marks': '0',
        'answer_explanation': 'Paris is the capital of France.',
        'question_order': '1', 'tags': 'geography',
    }
    out = io.StringIO()
    w = csv.DictWriter(out, fieldnames=ALL_COLS)
    w.writeheader()
    w.writerow({k: sample.get(k, '') for k in ALL_COLS})
    return out.getvalue()
